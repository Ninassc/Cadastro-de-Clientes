import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

ctk.set_appearance_mode('dark')
ctk.set_default_color_theme('dark-blue')

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout()
        self.aparencia()
        self.sistema()


    def layout(self):
        self.title('Cadastro de Clientes')
        self.geometry('850x500')

    def mudarAparencia(self, novaAparencia):
        ctk.set_appearance_mode(novaAparencia)

    def aparencia(self):
        self.label = ctk.CTkLabel(self, text='Tema', bg_color='transparent', text_color=['#000', '#fff'])
        self.label.place(x=20, y=430)

        self.opcoesAparencia = ctk.CTkOptionMenu(self, values=['Light', 'Dark', 'System'], command=self.mudarAparencia)
        self.opcoesAparencia.place(x=20, y=460)

    def sistema(self):
        frame = ctk.CTkFrame(self, width=850, height=50, corner_radius=0, fg_color='teal', bg_color='teal' )
        frame.place(x=0, y=10)

        titulo = ctk.CTkLabel(frame, text='Cadastro de Clientes', bg_color='teal', font=('Century Gothic bold', 24), text_color='#fff')
        titulo.place(y=15, x=320)

        spam = ctk.CTkLabel(self, text='Faça o Cadastro de Clientes!', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        spam.place(y=70, x=20)

        #Labels
        labelNome = ctk.CTkLabel(self, text='Nome Completo: ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelNome.place(y=120, x=20)

        labelContato = ctk.CTkLabel(self, text='Contato: ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelContato.place(y=200, x=20)

        labelIdade= ctk.CTkLabel(self, text='Idade: ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelIdade.place(y=200, x=500)

        labelGenero = ctk.CTkLabel(self, text='Gênero: ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelGenero.place(y=200, x=700)

        labelEndereco = ctk.CTkLabel(self, text='Endereço: ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelEndereco.place(y=120, x=500)

        labelObs = ctk.CTkLabel(self, text='Observações (não obrigatório): ', font=('Century Gothic bold', 14), text_color=['#000', '#fff'],)
        labelObs.place(y=280, x=20)

        #Variavels para text da entry (class não aceita o get() direto)
        nameValue = StringVar()
        contatoValue = StringVar()
        idadeValue = StringVar()
        enderecoValue = StringVar()
        
      

        #Entrys
        entryNome = ctk.CTkEntry(self, width=300, font=('Century Gothic', 12), fg_color='transparent', textvariable=nameValue)
        entryNome.place(y=150, x=20)

        entryContato = ctk.CTkEntry(self, width=250, font=('Century Gothic', 12), fg_color='transparent', textvariable=contatoValue)
        entryContato.place(y=230, x=20)

        entryIdade = ctk.CTkEntry(self, width=100, font=('Century Gothic', 12), fg_color='transparent', textvariable=idadeValue)
        entryIdade.place(y=230, x=500)

        entryEndereco = ctk.CTkEntry(self, width=300, font=('Century Gothic', 12), fg_color='transparent', textvariable=enderecoValue)
        entryEndereco.place(y=150, x=500)

        #caixa de texto
        tbObs = ctk.CTkTextbox(self, width=500, height=110, font=('Century Gothic', 12), fg_color='transparent', border_color='#aaa', border_width=2)
        tbObs.place(y=310, x=20)

        # Combobox
        cbGenero = ctk.CTkComboBox(self, values=['Masculino', 'Feminino'], font=('Century Gothic', 12),)
        cbGenero.set('Masculino')
        cbGenero.place(y=230, x=700)

        
        ficheiro = pathlib.Path('Clientes.xlsx')
        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = 'Nome Completo: '
            folha['B1'] = 'Contato: '
            folha['C1'] = 'Idade: '
            folha['D1'] = 'Endereço: '
            folha['E1'] = 'Gênero: '
            folha['F1'] = 'Obs: '

            ficheiro.save('Clientes.xlsx')

        #botãos
        def enviar():

            nome = entryNome.get()
            contato = entryContato.get()
            idade = entryIdade.get()
            endereco = entryEndereco.get()
            genero = cbGenero.get()
            obs = tbObs.get(0.0, END)

            if nome == "" or contato == "" or idade == "" or endereco == "":
                messagebox.showerror('Sistema', 'Erro! \nPreencha todos os campos! ')
            else: 

                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=nome)
                folha.cell(column=2, row=folha.max_row, value=contato)
                folha.cell(column=3, row=folha.max_row, value=idade)
                folha.cell(column=4, row=folha.max_row, value=endereco)
                folha.cell(column=5, row=folha.max_row, value=genero)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r'Clientes.xlsx')
                messagebox.showinfo('Sistema', 'Dados salvos com sucesso!')

        def limpar():
            nameValue.set("")
            contatoValue.set("")
            idadeValue.set("")
            enderecoValue.set("")
            tbObs.delete(0.0, END)

            messagebox.showinfo('Sistema', 'Tela limpa com sucesso!')


        botaoEnviar = ctk.CTkButton(self, text='Enviar', font=('Century Gothic bold', 14), command=enviar)
        botaoEnviar.place(x=500, y=460)

        botaoLimpar = ctk.CTkButton(self, text='Apagar', font=('Century Gothic bold', 14), command=limpar)
        botaoLimpar.place(x=650, y=460)


app = App()
app.mainloop()
